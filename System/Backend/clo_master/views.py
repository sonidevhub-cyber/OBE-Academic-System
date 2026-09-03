from collections import defaultdict

from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from assessments.models import CQI as CLOCQI
from core.models import Batch, Program, Semester
from retake.models import CourseRetake

from .models import CourseCLOMasterEntry, SemesterCLOMasterCache
from .serializers import SemesterCLOMasterCacheSerializer


class CLOMasterViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SemesterCLOMasterCache.objects.filter(is_active=True)
    serializer_class = SemesterCLOMasterCacheSerializer


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_clo_master_report(request, program_id, semester_id):
    """
    Get the master CLO compilation for a specific program and semester.
    Optional batch_id query param.
    Optional format=xlsx for Excel export.
    """
    # OPT-9: Clear per-batch scored_ids memo so this report load picks up
    # any marks / roster changes since the last request (no cross-request
    # staleness), while still sharing the sub-query within this request.
    from obe.services import clear_student_batch_caches
    clear_student_batch_caches()

    batch_id = request.query_params.get("batch_id")
    force_refresh = request.query_params.get("refresh") == "1"

    program = Program.objects.get(id=program_id)
    semester = Semester.objects.get(id=semester_id)
    batch = Batch.objects.get(id=batch_id) if batch_id else None

    # Get the curriculum version and valid courses for this batch & semester.
    # We intentionally use the curriculum as the source of truth so courses that
    # have been promoted but not finalized yet still appear in the report with
    # zeroed scores. We then ALSO append any enrolled Elective/Selective courses
    # because those are registered dynamically via StudentElectiveEnrollment
    # (they are NOT present in CurriculumVersionCourse).
    valid_course_ids = []
    course_catalog = []
    compulsory_ids_set = set()
    if batch and batch.curriculum_version:
        from curriculum.models import CurriculumVersionCourse
        curriculum_version_courses = CurriculumVersionCourse.objects.filter(
            version=batch.curriculum_version,
            semester_no=semester.number,
            is_active=True
        ).select_related('course').order_by('semester_no', 'course__code', 'course__name')
        valid_course_ids = [cvc.course.id for cvc in curriculum_version_courses]
        compulsory_ids_set = set(valid_course_ids)
        course_catalog = [
            {
                "course": cvc.course,
                "semester_no": cvc.semester_no,
            }
            for cvc in curriculum_version_courses
        ]

    # Merge active Elective/Selective courses that have sessions + enrolled
    # students for this (program, batch, semester) into the report columns.
    # Otherwise CS-2001 (AI elective) and similar courses never render a
    # column on the coordinator side even though instructor locked internals.
    from core.models.course import Course
    from obe.models import CourseSession as _CS
    elective_sessions_qs = _CS.objects.filter(
        course__program=program,
        semester=semester,
        is_active=True,
        course__offering_type__in=(Course.OFFERING_ELECTIVE, Course.OFFERING_SELECTIVE),
    ).select_related('course')
    if batch:
        elective_sessions_qs = elective_sessions_qs.filter(batch=batch)
    elective_sessions = list(elective_sessions_qs)
    seen_elective_ids = set()
    for es in elective_sessions:
        ecourse = es.course
        if ecourse.id in seen_elective_ids or ecourse.id in compulsory_ids_set:
            continue
        seen_elective_ids.add(ecourse.id)
        valid_course_ids.append(ecourse.id)
        course_catalog.append({
            "course": ecourse,
            "semester_no": semester.number,
        })

    # Keep the master cache aligned with live CLO scores on every report load,
    # similar to how GA reports read fresh CourseGAScore rows.
    master_cache = None
    if batch:
        from .signals import sync_stale_clo_master_cache

        latest_retake = (
            CourseRetake.objects.filter(current_batch=batch, is_active=True)
            .order_by("-updated_at")
            .first()
        )
        force_sync = force_refresh
        if (
            not force_sync
            and latest_retake
        ):
            existing_cache = SemesterCLOMasterCache.objects.filter(
                program=program,
                batch=batch,
                semester=semester,
            ).first()
            if (
                existing_cache is None
                or (
                    existing_cache.last_updated
                    and latest_retake.updated_at > existing_cache.last_updated
                )
            ):
                force_sync = True

        master_cache = sync_stale_clo_master_cache(
            program=program,
            batch=batch,
            semester=semester,
            valid_course_ids=valid_course_ids,
            force=force_sync,
        )

    # Get all course sessions for pending list.
    from obe.models import CourseSession

    all_course_sessions = CourseSession.objects.filter(
        course__program=program,
        semester=semester,
        is_active=True,
    )
    if batch_id:
        all_course_sessions = all_course_sessions.filter(batch_id=batch_id)
    if valid_course_ids:
        all_course_sessions = all_course_sessions.filter(course__id__in=valid_course_ids)
    pending_course_sessions = all_course_sessions.exclude(
        Q(assessment_status="ASSESSMENT_DONE")
        | Q(final_submitted=True)
        | Q(assessment_done=True)
    )

    sessions_by_course_id = {
        session.course_id: session
        for session in all_course_sessions.select_related("course", "instructor")
    }

    # Get all students for this batch.
    # Use get_students_for_batch so the student roster matches the one that
    # CLOService uses to build cache entries (includes students whose
    # Student.batch was set to this batch even if user.batch differs, as well
    # as frozen/retake students). This keeps total_students consistent with
    # kpi_achieved_lookup so attainment never exceeds 100%.
    from students.models import Student
    from obe.services import get_students_for_batch

    if batch_id and batch:
        students = get_students_for_batch(batch)
    else:
        students = Student.objects.none()

    from core.models.course import Course
    from obe.models import CLO as _CLOModel

    all_catalog_course_ids = [item["course"].id for item in course_catalog] if course_catalog else []
    curriculum_version_id = batch.curriculum_version_id if (batch and batch.curriculum_version) else None

    # --- OPT-1: Single batch-fetch of ALL CLOs for every catalog course
    # (replaces N per-course queries).  Curriculum-version filter for
    # compulsory applied in-memory below; electives skip it.
    all_clos_qs = _CLOModel.objects.filter(is_active=True)
    if all_catalog_course_ids:
        all_clos_qs = all_clos_qs.filter(course_id__in=all_catalog_course_ids)
    all_clos_raw = list(
        all_clos_qs.order_by("course_id", "order_number", "id")
    )
    clos_by_course_id = defaultdict(list)
    for clo in all_clos_raw:
        clos_by_course_id[clo.course_id].append(clo)

    # --- OPT-2/4: Batch-resolve enrolled student ids in ONE pass.
    # Compulsory -> reuse single get_students_for_batch result;
    # Elective/Selective -> ONE StudentElectiveEnrollment query over all
    # (course, batch, semester), then group by course_id.
    from obe.services import _get_enrolled_student_ids_for_course_session
    from electives.models import StudentElectiveEnrollment as _SEE
    from collections import defaultdict as _dd

    compulsory_student_str_ids = None
    if batch_id and batch:
        # Compute compulsory student UUID set once for entire request.
        # _get_enrolled_student_ids_for_course_session for a compulsory
        # session just returns get_students_for_batch; so precompute and
        # share across every compulsory course column + kpi denominator.
        _comp_ids = _get_enrolled_student_ids_for_course_session(
            # Use a fake "compulsory-shaped" session; the function only
            # reads .course.offering_type and .batch.  Since we already
            # confirmed it's a compulsory shape by using a non-elective
            # course id, shortcut directly.
            type(
                "_Sesh",
                (),
                {
                    "course": type(
                        "_C", (), {"offering_type": Course.OFFERING_COMPULSORY, "id": None}
                    )(),
                    "batch": batch,
                },
            )()
        )
        compulsory_student_str_ids = {str(sid) for sid in _comp_ids}

    elective_course_id_set = set()
    for course_item in course_catalog:
        _c = course_item["course"]
        if _c.offering_type in (Course.OFFERING_ELECTIVE, Course.OFFERING_SELECTIVE):
            elective_course_id_set.add(_c.id)

    elective_enrolled_by_course = {}  # course_id -> set(str_student_id)
    if elective_course_id_set and batch_id and batch:
        see_rows = _SEE.objects.filter(
            course_id__in=elective_course_id_set,
            batch=batch,
            semester=semester,
            is_active=True,
        ).values_list("course_id", "student_id")
        elective_enrolled_by_course = _dd(set)
        for cid, sid in see_rows:
            elective_enrolled_by_course[cid].add(str(sid))

    def _enrolled_ids_for_course(course_id, course):
        """Return enrolled str(student_id) set for the course, or None for compulsory/full-batch."""
        if course.offering_type in (Course.OFFERING_ELECTIVE, Course.OFFERING_SELECTIVE):
            return elective_enrolled_by_course.get(course_id, set())
        return None

    # Get all active course entries from cache (materialize ONCE as a list
    # so we can compute finalized_count and the lookup loops from the
    # same data without re-running a query).
    course_entries = []
    if master_cache:
        course_entries_query = CourseCLOMasterEntry.objects.filter(
            master_cache=master_cache,
            is_active=True
        ).select_related("student", "clo", "clo__course", "course_session")
        if valid_course_ids:
            course_entries_query = course_entries_query.filter(course__id__in=valid_course_ids)
        course_entries = list(course_entries_query)

    # --- OPT-3: Compute finalized_count DIRECTLY from the already-loaded
    # course_entries (avoids a redundant duplicate CourseCLOMasterEntry query
    # with values_list + distinct + count below).
    finalized_count = 0
    if master_cache and course_entries:
        _finalized_session_ids = set()
        for entry in course_entries:
            if entry.course_session_id:
                _finalized_session_ids.add(entry.course_session_id)
        finalized_count = len(_finalized_session_ids)

    # Build both enrollment maps in one pass:
    #   course_enrolled_ids_for_kpi  (for the cache-entry loop below)
    #   course_enrolled_student_ids  (for sorted_courses rendering)
    course_enrolled_ids_for_kpi = {}
    course_enrolled_student_ids = {}

    # First, seed course_enrolled_ids_for_kpi from sessions_by_course_id
    # entries that are elective/selective (mirrors prior semantics, but now
    # uses the batched elective_enrolled_by_course map).
    for session in all_course_sessions.select_related("course"):
        _c = session.course
        _ids = _enrolled_ids_for_course(_c.id, _c)
        if _ids is not None:
            course_enrolled_ids_for_kpi[_c.id] = _ids

    # Precompute lookups for O(1) access.
    course_entry_lookup = {}
    kpi_achieved_lookup = {}
    kpi_total_lookup = {}  # Count total students per CLO (from cache entries)

    for entry in course_entries:
        student_id = entry.student.student_id
        course_id = entry.course.id
        clo_id = entry.clo.id

        enrolled_ids = course_enrolled_ids_for_kpi.get(course_id)
        if enrolled_ids is not None and str(student_id) not in enrolled_ids:
            continue

        course_entry_lookup[(student_id, course_id, clo_id)] = entry

        kpi_key = (course_id, clo_id)
        kpi_achieved_lookup.setdefault(kpi_key, 0)
        kpi_total_lookup.setdefault(kpi_key, 0)
        kpi_total_lookup[kpi_key] += 1  # Count total students for this CLO
        if entry.is_kpi_achieved:
            kpi_achieved_lookup[kpi_key] += 1

    # Fetch all approved CQIs for this program, batch, semester.
    cqi_filter = {
        "course__program": program,
        "semester": semester,
        "status": "approved",
    }
    if batch:
        cqi_filter["batch"] = batch
    approved_cqis = CLOCQI.objects.filter(**cqi_filter).select_related("course", "clo")

    cqi_lookup = {}
    for cqi in approved_cqis:
        cqi_lookup[(cqi.course.id, cqi.clo.id)] = cqi

    # Build course -> CLO catalog from the curriculum so planned but not yet
    # finalized courses still render.
    if not course_catalog and course_entries:
        # Fallback for legacy batches without a curriculum mapping.
        distinct_courses = {}
        for entry in course_entries:
            distinct_courses[entry.course.id] = entry.course
        course_catalog = [
            {"course": course, "semester_no": semester.number}
            for course in sorted(distinct_courses.values(), key=lambda c: c.code)
        ]
        # Re-run OPT-1/2 setup since course_catalog changed.
        all_catalog_course_ids = [item["course"].id for item in course_catalog]
        if all_catalog_course_ids and not all_clos_raw:
            all_clos_raw = list(
                _CLOModel.objects.filter(
                    is_active=True, course_id__in=all_catalog_course_ids
                ).order_by("course_id", "order_number", "id")
            )
            clos_by_course_id = defaultdict(list)
            for clo in all_clos_raw:
                clos_by_course_id[clo.course_id].append(clo)

    sorted_courses = []
    for course_item in course_catalog:
        course = course_item["course"]
        course_id = course.id

        # OPT-1: Use pre-fetched CLO list filtered + ordered in memory.
        raw_clos = clos_by_course_id.get(course_id, [])
        use_curriculum_filter = bool(
            curriculum_version_id
            and course.offering_type not in (Course.OFFERING_ELECTIVE, Course.OFFERING_SELECTIVE)
        )
        if use_curriculum_filter:
            filtered_clos = [
                c for c in raw_clos if c.curriculum_version_id == curriculum_version_id
            ]
        else:
            filtered_clos = list(raw_clos)
        # filtered_clos already ordered by (order_number, id) from the
        # master order_by; no extra sort needed.

        sorted_courses.append({
            "course": course,
            "semester_no": course_item["semester_no"],
            "clos": filtered_clos,
        })

        # OPT-2/4: Build course_enrolled_student_ids in the SAME loop
        # (eliminates second sorted_courses iteration + function calls).
        _eids = _enrolled_ids_for_course(course_id, course)
        course_enrolled_student_ids[course_id] = _eids

    total_students = len(compulsory_student_str_ids) if (compulsory_student_str_ids is not None) else students.count()
    course_cohort_totals = {}
    for course_info in sorted_courses:
        course_id = course_info["course"].id
        enrolled_ids = course_enrolled_student_ids.get(course_id)
        if enrolled_ids is not None:
            course_cohort_totals[course_id] = len(enrolled_ids)
        else:
            course_cohort_totals[course_id] = total_students

    # Prepare student data.
    students_data = []
    for student in students:
        row = {
            "sr_no": len(students_data) + 1,
            "reg_no": getattr(student, "registration_number", ""),
            "name": getattr(student, "name", student.user.full_name),
            "courses": {},
        }

        for course_info in sorted_courses:
            course_id = course_info["course"].id
            course_key = str(course_id)
            enrolled_ids = course_enrolled_student_ids.get(course_id)
            is_enrolled = enrolled_ids is None or student.student_id in enrolled_ids
            row["courses"][course_key] = {}

            for clo in course_info["clos"]:
                clo_key = f"CLO-{clo.order_number}"
                if not is_enrolled:
                    row["courses"][course_key][clo_key] = None
                else:
                    entry_obj = course_entry_lookup.get((student.student_id, course_id, clo.id))
                    if entry_obj:
                        row["courses"][course_key][clo_key] = {
                            "score": float(entry_obj.clo_score),
                            "achieved": entry_obj.is_kpi_achieved,
                        }
                    else:
                        row["courses"][course_key][clo_key] = None

        students_data.append(row)

    kpi_achieved_counts = {}
    for course_info in sorted_courses:
        for clo in course_info["clos"]:
            clo_key = f"{course_info['course'].id}-CLO-{clo.order_number}"
            achieved_count = kpi_achieved_lookup.get((course_info["course"].id, clo.id), 0)
            kpi_achieved_counts[clo_key] = achieved_count

    pending_courses_info = []
    for cs in pending_course_sessions:
        if getattr(cs, "internal_complete_awaiting_final", False):
            pending_status = "FINAL_SUBMISSION_PENDING"
        else:
            pending_status = "INTERNAL_LOCK_PENDING"
        pending_courses_info.append({
            "course_id": cs.course.id,
            "course_code": cs.course.code,
            "course_name": cs.course.name,
            "instructor_name": cs.instructor.full_name if cs.instructor else "Not Assigned",
            "status": pending_status,
        })

    all_session_course_ids = {cs.course.id for cs in all_course_sessions}
    for course_info in sorted_courses:
        course = course_info["course"]
        if course.id in all_session_course_ids:
            continue
        pending_courses_info.append(
            {
                "course_id": course.id,
                "course_code": course.code,
                "course_name": course.name,
                "instructor_name": "Not Assigned",
                "status": "NO_SESSION_CREATED",
            }
        )

    total_count = len(sorted_courses) if sorted_courses else all_course_sessions.count()
    is_fully_compiled = finalized_count >= total_count if total_count else False

    response_data = {
        "program": {
            "id": program.id,
            "name": program.name,
            "code": program.code,
        },
        "semester": {
            "id": semester.id,
            "name": semester.name,
            "number": semester.number,
        },
        "batch": {
            "id": batch_id,
            "name": batch.name if batch else "",
        }
        if batch_id
        else None,
        "status": {
            "finalized_count": finalized_count,
            "total_count": total_count,
            "is_fully_compiled": is_fully_compiled,
        },
        "finalized_courses": [
            {
                "course_id": info["course"].id,
                "course_code": info["course"].code,
                "course_name": info["course"].name,
                "clos": [
                    {
                        "clo_id": clo.id,
                        "clo_code": f"CLO-{clo.order_number}",
                        "clo_name": clo.title,
                        "kpi_target": clo.kpi_target,
                        "cohort_achieved_count": kpi_achieved_lookup.get(
                            (info["course"].id, clo.id), 0
                        ),
                         "cohort_total_count": course_cohort_totals.get(info["course"].id, total_students),
                         "cohort_percentage": (
                             (
                                 kpi_achieved_lookup.get((info["course"].id, clo.id), 0)
                                 / course_cohort_totals.get(info["course"].id, total_students)
                                 * 100
                             )
                             if course_cohort_totals.get(info["course"].id, total_students) > 0
                             else 0
                        ),
                        "cqi": {
                            "reason": cqi_lookup[(info["course"].id, clo.id)].reason,
                            "action_plan": cqi_lookup[(info["course"].id, clo.id)].action_plan,
                            "coordinator_comment": cqi_lookup[(info["course"].id, clo.id)].coordinator_comment,
                        }
                        if (info["course"].id, clo.id) in cqi_lookup
                        else None,
                    }
                    for clo in info["clos"]
                ],
            }
            for info in sorted_courses
        ],
        "students": students_data,
        "pending_courses": pending_courses_info,
        "summary": {
            "total_students": total_students,
            "kpi_breakdown": kpi_achieved_counts,
        },
    }

    is_export_route = bool(request.resolver_match and request.resolver_match.url_name == "clo-master-report-export")
    is_xlsx_export = request.query_params.get("format") == "xlsx" or is_export_route

    if is_xlsx_export:
        return export_to_excel(response_data)

    return Response(response_data)


def export_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "CLO Master Compilation"

    header_font = Font(name="Arial", size=12, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    title = f"CLO Master Compilation - {data['program']['name']} - {data['semester']['name']}"
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws.merge_cells("A1:Z1")
    ws["A1"].alignment = header_alignment

    ws["A3"] = f"Courses Finalized: {data['status']['finalized_count']}/{data['status']['total_count']}"
    ws["A3"].font = Font(bold=True)
    ws.merge_cells("A3:Z3")

    headers = ["Sr. No", "Reg. No", "Student Name"]
    for course in data["finalized_courses"]:
        for clo in course["clos"]:
            cohort_pct = round(float(clo.get("cohort_percentage", 0)), 2)
            headers.append(
                f"{course['course_code']} - {clo['clo_code']} (Target: {clo.get('kpi_target', 0)}%, Cohort: {cohort_pct:.2f}%)"
            )
            if clo.get("cqi"):
                headers.append(f"{course['course_code']} - {clo['clo_code']} - CQI Reason")
                headers.append(f"{course['course_code']} - {clo['clo_code']} - CQI Action Plan")
                if clo["cqi"].get("coordinator_comment"):
                    headers.append(
                        f"{course['course_code']} - {clo['clo_code']} - Coordinator Comment"
                    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    last_student_row = 5
    for row_num, student in enumerate(data["students"], 6):
        last_student_row = row_num
        ws.cell(row=row_num, column=1, value=student["sr_no"]).alignment = cell_alignment
        ws.cell(row=row_num, column=2, value=student["reg_no"]).alignment = cell_alignment
        ws.cell(row=row_num, column=3, value=student["name"]).alignment = cell_alignment

        col = 4
        for course in data["finalized_courses"]:
            course_id = str(course["course_id"])
            for clo in course["clos"]:
                clo_code = clo["clo_code"]
                course_data = student["courses"].get(course_id, {})
                clo_data = course_data.get(clo_code)

                if clo_data is not None:
                    score_val = round(float(clo_data["score"]), 1)
                    cell_val = f"{score_val:.1f}%"
                    cell = ws.cell(row=row_num, column=col, value=cell_val)
                    cell.alignment = cell_alignment
                    if clo_data["achieved"]:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    ws.cell(row=row_num, column=col, value="-").alignment = cell_alignment
                col += 1

                if clo.get("cqi"):
                    ws.cell(row=row_num, column=col, value=clo["cqi"]["reason"]).alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                    col += 1
                    ws.cell(row=row_num, column=col, value=clo["cqi"]["action_plan"]).alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                    col += 1
                    if clo["cqi"].get("coordinator_comment"):
                        ws.cell(
                            row=row_num,
                            column=col,
                            value=clo["cqi"]["coordinator_comment"],
                        ).alignment = Alignment(horizontal="left", vertical="center")
                        col += 1

    total_students = data["summary"]["total_students"]
    summary_row1 = last_student_row + 1
    summary_row2 = last_student_row + 2

    ws.cell(
        row=summary_row1,
        column=1,
        value="No. of Students Achieving CLOs KPI (50%):",
    )
    ws.merge_cells(start_row=summary_row1, start_column=1, end_row=summary_row1, end_column=3)
    ws.cell(row=summary_row1, column=1).font = Font(name="Arial", size=12, bold=True)
    ws.cell(row=summary_row1, column=1).alignment = cell_alignment

    ws.cell(
        row=summary_row2,
        column=1,
        value="% of Students Achieving CLOs at Cohort-Level (50%):",
    )
    ws.merge_cells(start_row=summary_row2, start_column=1, end_row=summary_row2, end_column=3)
    ws.cell(row=summary_row2, column=1).font = Font(name="Arial", size=12, bold=True)
    ws.cell(row=summary_row2, column=1).alignment = cell_alignment

    col = 4
    for course in data["finalized_courses"]:
        course_id = str(course["course_id"])
        for clo in course["clos"]:
            achieved_count = int(clo.get("cohort_achieved_count", 0))

            cell1 = ws.cell(row=summary_row1, column=col, value=achieved_count)
            cell1.font = Font(bold=True)
            cell1.alignment = cell_alignment

            cohort_pct = round(float(clo.get("cohort_percentage", 0)), 2)
            percentage = cohort_pct
            cell2 = ws.cell(row=summary_row2, column=col, value=f"{percentage:.2f}%")
            cell2.font = Font(bold=True)
            cell2.alignment = cell_alignment
            col += 1

            if clo.get("cqi"):
                col += 1
                col += 1
                if clo["cqi"].get("coordinator_comment"):
                    col += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = (max_length + 2) * 1.5
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="CLO_Master_Compilation_{data["program"]["code"]}_{data["semester"]["name"]}.xlsx"'
    )
    wb.save(response)
    return response
